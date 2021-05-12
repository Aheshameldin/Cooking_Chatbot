import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook


def ing():
    url = input("Enter URL: ")
    try:
        r = requests.get(url, timeout=30)
    except requests.ConnectionError as e:
        print("OOPS!! Connection Error. Make sure you are connected to Internet. Technical Details given below.\n")
        print(str(e))
    except requests.Timeout as e:
        print("OOPS!! Timeout Error")
        print(str(e))
    except requests.RequestException as e:
        print("OOPS!! General Error")
        print(str(e))
    except KeyboardInterrupt:
        print("Someone closed the program")


    soup = BeautifulSoup(r.text, 'html.parser')


    name = soup.find('span', {'id': 'hs_cos_wrapper_name'}).text
    # print(name)
    # print(len(name))


    ingredients=[]
    ing = soup.find_all('li', {'class': 'ml-4 pl-2'})
    for x in ing:
        ingredients.append(x.text)
    # print(ingredients)


    steps=[]
    list_of_steps = soup.find_all('h3')
    for i in list_of_steps:
        steps.append(i.text)
    # print(steps)

    details=[]
    list_of_details = soup.find('ol',{'class': 'mb-0 instructions-list'})
    detail= list_of_details.find_all('li',{'class': 'mb-2'})
    for r in detail:
        details.append(r.text)
    # print(details)

    output_dict={'Ingredients': pd.Series(ingredients),
                     'Steps':pd.Series(steps),
                     'Details':pd.Series(details),
                     'URL': url}

    # path= 'Cooking.xlsx'
    book = load_workbook('Cooking.xlsx')
    writer = pd.ExcelWriter('Cooking.xlsx', engine='openpyxl')
    writer.book = book
    df = pd.DataFrame(output_dict)
    df.to_excel(writer, sheet_name = name[:30])
    writer.save()
    writer.close()
    # sheet = book.create_sheet(name[:30])
    # df.to_excel('Cooking.xlsx', engine='xlsxwriter', index=False)
    # sheet.append(data)
    print("Done")
    # book.save('Cooking.xlsx')

# ing()


#To separate each sheet into a separate csv
# loop through the dictionary and save csv
data = pd.read_excel('Cooking.xlsx', sheet_name=None)
for sheet_name, df in data.items():
    df.to_csv(f'{sheet_name}.csv')

def cooking():
    no_of_yes=0
    shopping_list=[]
    xl = pd.ExcelFile('Cooking.xlsx')
    menu= []
    for x in xl.sheet_names:
        menu.append(x)
    print("What do you want to cook?")
    print("I have the following:", *menu, sep=",")
    userInput = input(">>>You:")


    if userInput in xl.sheet_names:
        print("I will search for ingredients")
        file=userInput
        df = pd.read_csv(f"{file}.csv")
        ing_column = df["Ingredients"]  # you can also use df['column_name']
        # print(len(ing_column))
        print("Do you want to skip List?")
        userInput = input(">>>You:")
        if userInput.lower() =="no":
            for x in ing_column:
                print("Do you have", x, "?")
                reply = input(">>>You:")
                while True:
                    if reply == "yes":
                        no_of_yes+=1
                        break
                    elif reply=="no":
                        shopping_list.append(x)
                        print("Added to Shopping list")
                        break
                    else:
                        print("I dont understand yes or no??")
                        reply = input(">>>You:")
                        if reply == "yes":
                            continue
                        elif reply == "no":
                            shopping_list.append(x)
                            print("Added to Shopping list")
        else:
            no_of_yes = len(ing_column)

        if no_of_yes==len(ing_column):
            print("Here are the steps")
            steps = df["Details"].dropna().unique()
            count=0
            for x in steps:
                count+=1
                print(count,x)
            print("Bonne Appetite")

        else:
            print("go buy", *shopping_list,sep=",")
cooking()



